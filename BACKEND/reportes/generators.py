# reportes/generators.py
import io
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # <-- 1. IMPORTADO ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch  # <-- 2. IMPORTADO inch
import openpyxl
from openpyxl.styles import Font, Alignment
from decimal import Decimal
from datetime import datetime, date

# --- GENERADOR DE PDF (con ReportLab) ---
def generar_reporte_pdf(data_para_reporte, interpretacion):
    """
    Toma una lista de diccionarios (datos) y la interpretación (para el título)
    y devuelve una respuesta HttpResponse con el PDF.
    """
    buffer = io.BytesIO()
    
    # Usar 'landscape' (horizontal) si hay muchas columnas
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    styles = getSampleStyleSheet()

    # 1. Título (basado en el prompt o interpretación)
    prompt_usado = interpretacion.get('prompt', 'Reporte Personalizado') # Necesitarías pasar el prompt original
    titulo = Paragraph(f"Reporte: {prompt_usado}", styles['h1'])
    titulo.style.alignment = 1 # Centrado
    story.append(titulo)
    story.append(Spacer(1, 12)) # Espacio

    # 2. Preparar Datos de la Tabla
    if not data_para_reporte:
        story.append(Paragraph("No se encontraron datos.", styles['Normal']))
    else:
        # Extraer cabeceras (headers) y filas (rows)
        headers = list(data_para_reporte[0].keys())
        # Formatear cabeceras (ej. 'cliente__nombre' -> 'Cliente')
        formatted_headers = [h.split('__')[-1].replace('_', ' ').capitalize() for h in headers]
        
        # Convertir todos los datos a string para la tabla
        rows = [formatted_headers] # Primera fila son las cabeceras
        for item in data_para_reporte:
            row_data = [str(item.get(h, '')) for h in headers] # Obtener valor y convertir a string
            rows.append(row_data)

        # 3. Crear Tabla y Estilos
        table = Table(rows)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4A90E2")), # Color azul para cabecera
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f0f4f8")), # Color fila alterna
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.darkgrey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ])
        table.setStyle(style)
        
        # Aplicar estilo de fila alterna
        for i, row in enumerate(rows):
            if i % 2 == 0 and i > 0: # Omitir cabecera
                style.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)

        story.append(table)

    # 4. Construir el PDF
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    # Usamos el prompt para un nombre de archivo dinámico
    filename = f"reporte_{prompt_usado.lower().replace(' ', '_')[:30]}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# --- GENERADOR DE EXCEL (con OpenPyXL) ---
def generar_reporte_excel(data_para_reporte, interpretacion):
    """
    Toma una lista de diccionarios (datos) y devuelve una respuesta HttpResponse
    con el archivo Excel (.xlsx).
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    prompt_usado = interpretacion.get('prompt', 'Reporte')
    sheet.title = prompt_usado[:30] # Límite de 31 chars para título de hoja

    if not data_para_reporte:
        sheet.cell(row=1, column=1).value = "No se encontraron datos."
    else:
        # 1. Escribir Cabeceras
        headers = list(data_para_reporte[0].keys())
        formatted_headers = [h.split('__')[-1].replace('_', ' ').capitalize() for h in headers]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_num, header_title in enumerate(formatted_headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        # 2. Escribir Datos
        for row_num, row_data in enumerate(data_para_reporte, 2): # Empezar en fila 2
            for col_num, header_key in enumerate(headers, 1):
                value = row_data.get(header_key, None)
                cell = sheet.cell(row=row_num, column=col_num)
                
                # Convertir tipos (Decimal, Datetime) a tipos que Excel entiende
                if isinstance(value, Decimal):
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                elif isinstance(value, datetime):
                    cell.value = value.replace(tzinfo=None) # Quitar timezone info
                    cell.number_format = 'YYYY-MM-DD HH:MM'
                elif isinstance(value, date):
                    cell.value = value
                    cell.number_format = 'YYYY-MM-DD'
                elif isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value) if value is not None else ''

    # 3. Guardar en Buffer de Memoria
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"reporte_{prompt_usado.lower().replace(' ', '_')[:30]}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# --- GENERADOR DE COMPROBANTE ---
def generar_comprobante_pdf(venta):
    """
    Toma un objeto Venta y genera su comprobante en PDF.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch/2, bottomMargin=inch/2)
    story = []
    styles = getSampleStyleSheet()

    # Estilo para texto normal
    normal_style = styles['Normal']
    # Estilo para texto en negrita
    bold_style = ParagraphStyle(name='Bold', parent=styles['Normal'], fontName='Helvetica-Bold')

    # 1. Título
    story.append(Paragraph("SmartSales365 - Comprobante de Venta", styles['h1']))
    story.append(Spacer(1, 0.25 * inch))

    # 2. Información de la Venta y Cliente
    info_table_data = [
        [Paragraph('Número de Venta:', bold_style), Paragraph(f"#{venta.id}", normal_style)],
        [Paragraph('Fecha:', bold_style), Paragraph(venta.fecha_venta.strftime('%d/%m/%Y %H:%M:%S'), normal_style)],
    ]
    if venta.cliente:
        info_table_data.append([Paragraph('Cliente:', bold_style), Paragraph(f"{venta.cliente.nombre} {venta.cliente.apellido}", normal_style)])
        info_table_data.append([Paragraph('Email Cliente:', bold_style), Paragraph(venta.cliente.email, normal_style)])
        if venta.cliente.direccion:
            info_table_data.append([Paragraph('Dirección:', bold_style), Paragraph(venta.cliente.direccion, normal_style)])
    if venta.vendedor:
        info_table_data.append([Paragraph('Atendido por:', bold_style), Paragraph(venta.vendedor.username, normal_style)])
    
    info_table = Table(info_table_data, colWidths=[1.5 * inch, 5.5 * inch]) # Ajustado ancho
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), # Quita o pon bordes
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.25 * inch))

    # 3. Detalles de la Venta (Productos)
    story.append(Paragraph("Detalles de la Compra:", styles['h3']))
    
    detalles_data = [
        # Cabeceras de la tabla
        [Paragraph('Producto', bold_style), Paragraph('Cant.', bold_style), Paragraph('P. Unit.', bold_style), Paragraph('Subtotal', bold_style)]
    ]
    
    # --- 3. OPTIMIZACIÓN (N+1 Query) ---
    # Usamos select_related('producto') para evitar una consulta por cada item
    detalles = venta.detalles.select_related('producto').all()
    # --- FIN OPTIMIZACIÓN ---
    
    if not detalles:
         detalles_data.append(["No se encontraron detalles", "", "", ""])
    else:
        for detalle in detalles:
            detalles_data.append([
                Paragraph(detalle.producto.nombre, normal_style), # Nombre del producto
                Paragraph(str(detalle.cantidad), normal_style),
                Paragraph(f"Bs {detalle.precio_unitario:.2f}", normal_style),
                Paragraph(f"Bs {detalle.subtotal:.2f}", normal_style)
            ])

    # Estilo de la tabla de detalles
    detalles_table = Table(detalles_data, colWidths=[3.5 * inch, 0.5 * inch, 1.5 * inch, 1.5 * inch]) # Ajustado ancho
    detalles_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4A90E2")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'), # Alinear números a la derecha
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(detalles_table)
    story.append(Spacer(1, 0.2 * inch))

    # 4. Total
    total_table_data = [
        [Paragraph('TOTAL VENTA:', bold_style), Paragraph(f"Bs {venta.total:.2f}", bold_style)]
    ]
    total_table = Table(total_table_data, colWidths=[5 * inch, 2 * inch])
    total_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
    ]))
    story.append(total_table)

    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    
    # Devolver la respuesta HTTP
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="comprobante_venta_{venta.id}.pdf"'
    return response